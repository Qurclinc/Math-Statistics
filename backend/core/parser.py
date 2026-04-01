from pathlib import Path
from openpyxl import load_workbook, Workbook

class Parser():
    
    def __init__(
        self,
        workbook: str | None = None,
        skip_cols: int = 1,
        sign_col: int = 0
    ):
        self._workbook = None
        self._sheet = None
        if workbook:
            self.workbook = workbook
        self.header = []
        self.signs = []
        self.data = []
        self._skip_cols = skip_cols
        self._sign_col = sign_col
        
    @property
    def workbook(self) -> Workbook:
        return self._workbook
    
    @property
    def sheet(self):
        return self._sheet
        
    @workbook.setter
    def workbook(self, workbook):
        self._workbook = load_workbook(workbook)
        self._sheet = self.workbook.worksheets[0]
        
    @property
    def skip_cols(self):
        return self._skip_cols
    
    @skip_cols.setter
    def skip_cols(self, skip_cols):
        if skip_cols < 0:
            raise ValueError
        self._skip_cols = skip_cols
    
    @property
    def sign_col(self):
        return self._sign_col
        
    @sign_col.setter
    def sign_col(self, sign_col):
        if sign_col < 1:
            raise ValueError
        self._sign_col = sign_col - 1
        
    def parse_data(self):
        self.header.clear()
        self.signs.clear()
        self.data.clear()
        for i, row in enumerate(self.sheet.iter_rows(values_only=True)):
            line = []
            if any(row):
                for j, value in enumerate(row):
                    if value is not None:
                        value = str(value).replace("\n", " ")
                        if j < self.skip_cols:
                            if j == self.sign_col:
                                self.signs.append(value) if i != 0 else self.header.append(value)
                            continue
                        if i == 0:
                            self.header.append(value)
                        else:
                            line.append(float(value))
                if line:
                    self.data.append(line)
                    
                    
parser = Parser()