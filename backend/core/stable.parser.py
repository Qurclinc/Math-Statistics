from pathlib import Path
from openpyxl import load_workbook

class Parser():
    
    def __init__(
        self,
        workbook: Path | str,
        skip_cols: int = 3,
        sign_col: int = 1
    ):
        self.wb = load_workbook(workbook)
        self.sheet = self.wb.worksheets[0]
        self.header = []
        self.signs = []
        self.data = []
        self.skip_cols = skip_cols
        self.sign_col = sign_col
        self.__parse_data()
        
    def __parse_data(self):
        for i, row in enumerate(self.sheet.iter_rows(values_only=True)):
            line = []
            if any(row):
                for j, value in enumerate(row):
                    if value:
                        value = str(value).replace("\n", " ")
                        if j < self.skip_cols:
                            if j == self.sign_col:
                                self.signs.append(value) if i != 0 else self.header.append(value)
                            continue
                        if i == 0:
                            self.header.append(value)
                        else:
                            line.append(float(value))
                if line:
                    self.data.append(line)